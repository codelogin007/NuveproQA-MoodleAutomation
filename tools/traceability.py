"""
BDD traceability & execution report for the Java/Cucumber suite.

Merges three sources into one per-case report:
  1) Docs/Testcases/MoodleRegression_Testing.xlsx  -> ALL manual cases (denominator)
  2) src/test/resources/features/*.feature          -> scenario @tags = CREATED/automated
  3) target/cucumber.json                            -> EXECUTED scenarios + pass/fail/skip

The @<AREA>-<n> tag on each scenario is the link between a manual case and its script.

Run (after `mvn test` has produced target/cucumber.json):
    <python-with-openpyxl> tools/traceability.py
Outputs (in reports/):
    traceability_execution.csv
    traceability_execution.html
"""
from __future__ import annotations
import csv
import html
import json
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

JAVA_ROOT = Path(__file__).resolve().parents[1]
FEATURES_DIR = JAVA_ROOT / "src" / "test" / "resources" / "features"
TARGET = JAVA_ROOT / "target"
# Merge results from every saved run (fast + slow) so the report reflects all executions.
CUKE_JSONS = ["cucumber_fast.json", "cucumber_slow.json", "cucumber.json"]
OUT_DIR = JAVA_ROOT / "reports"
XLSX = Path(r"D:\Projects\Nuvepro\Moodle\Docs\Testcases\MoodleRegression_Testing.xlsx")
GENERATED = "2026-07-01"

SHEET_PREFIX = {
    "Moodle_users": "USR", "Moodle_Assessment_Labs": "ASMT", "Moodle_Playground_Labs": "PG",
    "Moodle_Guided_Labs": "GP", "Moodle_Lab_Templates": "LT", "Groups_Testcases": "GRP",
    "Moodle_Roles": "ROLE", "Moodle_Tags": "TAG", "Moodle_Logs": "LOG",
    "LoadBalacer_Moodle": "LB", "Moodle_Reports": "RPT", "Moodle_cron": "CRON",
    "Moodle-Cloudlabs Integration": "INTG", "All Course Page ": "ACP",
    "Moodle settings": "SET", "Activities": "ACT",
}
AREA_NAME = {v: k.strip() for k, v in SHEET_PREFIX.items()}
ID_RE = re.compile(r"([A-Z]{2,5}-\d+[a-z]?)", re.IGNORECASE)


def base(cid: str) -> str:
    return re.sub(r"^([A-Z]{2,5}-\d+)[A-Za-z]$", r"\1", cid.upper())


def _norm(h):
    return str(h).strip().lower() if h is not None else ""


def _col(headers, *names):
    low = [_norm(h) for h in headers]
    for n in names:
        for i, h in enumerate(low):
            if h == n.lower():
                return i
    for n in names:
        for i, h in enumerate(low):
            if n.lower() in h:
                return i
    return None


def read_manual_cases():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    cases = []
    for sheet, area in SHEET_PREFIX.items():
        if sheet not in wb.sheetnames:
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        headers = rows[0]
        c_sno = _col(headers, "s.no", "sl.no", "s. no")
        c_title = _col(headers, "summary", "test cases", "testcases", "name")
        c_imp = _col(headers, "importance")

        def title_of(r):
            v = r[c_title] if c_title is not None and c_title < len(r) else None
            return str(v).strip() if v is not None else ""

        def sno_of(r):
            return r[c_sno] if c_sno is not None and c_sno < len(r) else None

        title_rows = [r for r in rows[1:] if title_of(r)]
        numbered = sum(1 for r in title_rows if isinstance(sno_of(r), (int, float)))
        use_sno = title_rows and numbered >= max(1, len(title_rows) // 2)
        running = 0
        for r in title_rows:
            running += 1
            sno = sno_of(r)
            num = int(sno) if (use_sno and isinstance(sno, (int, float))) else running
            cases.append({
                "id": f"{area}-{num}", "area": area, "num": num,
                "title": title_of(r)[:140],
                "importance": (str(r[c_imp]).strip() if c_imp is not None and c_imp < len(r) and r[c_imp] else ""),
            })
    wb.close()
    return cases


def scan_features():
    """{base_id: {'scenario': name, 'feature': filename}} from @tags on scenarios."""
    created = {}
    if not FEATURES_DIR.exists():
        return created
    for f in sorted(FEATURES_DIR.glob("*.feature")):
        pending = []
        for line in f.read_text(encoding="utf-8", errors="ignore").splitlines():
            s = line.strip()
            if s.startswith("@"):
                pending += [t for t in ID_RE.findall(s)]
            elif s.lower().startswith("scenario"):
                name = s.split(":", 1)[1].strip() if ":" in s else s
                for t in pending:
                    created[base(t)] = {"scenario": name, "feature": f.name}
                pending = []
            elif s.lower().startswith("feature"):
                pending = []
    return created


def read_execution():
    """{base_id: 'passed'|'failed'|'skipped'} merged across all saved cucumber json runs."""
    out = {}
    for name in CUKE_JSONS:
        p = TARGET / name
        if not p.exists():
            continue
        data = json.loads(p.read_text(encoding="utf-8"))
        for feat in data:
            for el in feat.get("elements", []):
                if el.get("type") != "scenario":
                    continue
                statuses = [st.get("result", {}).get("status") for st in el.get("steps", [])]
                if any(s in ("failed", "undefined", "pending", "ambiguous") for s in statuses):
                    outcome = "failed"
                elif any(s == "skipped" for s in statuses):
                    outcome = "skipped"
                else:
                    outcome = "passed"
                for tag in el.get("tags", []):
                    tname = tag.get("name", "").lstrip("@")
                    if ID_RE.fullmatch(tname):
                        out[base(tname)] = outcome
    return out


def main():
    OUT_DIR.mkdir(exist_ok=True)
    cases = read_manual_cases()
    created = scan_features()
    executed = read_execution()

    for c in cases:
        d = created.get(c["id"])
        c["automated"] = "yes" if d else "no"
        c["scenario"] = d["scenario"] if d else ""
        c["feature"] = d["feature"] if d else ""
        c["result"] = executed.get(c["id"], "")
        c["executed"] = "yes" if c["id"] in executed else "no"

    # CSV
    with (OUT_DIR / "traceability_execution.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["id", "area", "importance", "automated",
                                          "executed", "result", "scenario", "feature", "title"],
                           extrasaction="ignore")
        w.writeheader()
        for c in sorted(cases, key=lambda x: (x["area"], x["num"])):
            w.writerow(c)

    per = defaultdict(lambda: {"total": 0, "auto": 0, "exec": 0, "pass": 0})
    for c in cases:
        a = per[c["area"]]
        a["total"] += 1
        if c["automated"] == "yes":
            a["auto"] += 1
        if c["executed"] == "yes":
            a["exec"] += 1
        if c["result"] == "passed":
            a["pass"] += 1
    tot = sum(a["total"] for a in per.values())
    auto = sum(a["auto"] for a in per.values())
    ex = sum(a["exec"] for a in per.values())
    ps = sum(a["pass"] for a in per.values())

    color = {"passed": "#1a7f37", "failed": "#cf222e", "skipped": "#9a6700", "": "#8b949e"}
    label = {"passed": "PASS", "failed": "FAIL", "skipped": "SKIP", "": "not run"}

    def esc(s):
        return html.escape(str(s))

    summary = "".join(
        f"<tr><td>{a}</td><td>{esc(AREA_NAME.get(a,''))}</td><td class=r>{per[a]['auto']}</td>"
        f"<td class=r>{per[a]['exec']}</td><td class=r>{per[a]['pass']}</td><td class=r>{per[a]['total']}</td>"
        f"<td class=r>{(per[a]['auto']/per[a]['total']*100 if per[a]['total'] else 0):.0f}%</td></tr>"
        for a in sorted(per))
    detail = "".join(
        f"<tr><td><b>{esc(c['id'])}</b></td><td>{esc(c['title'])}</td>"
        f"<td>{'yes' if c['automated']=='yes' else '—'}</td>"
        f"<td>{esc(c['scenario'])}</td>"
        f"<td><span class=badge style='background:{color[c['result']]}'>{label[c['result']]}</span></td></tr>"
        for c in sorted((c for c in cases if c["automated"] == "yes"), key=lambda x: (x["area"], x["num"])))

    doc = f"""<!doctype html><html><head><meta charset=utf-8><title>Traceability &amp; Execution</title>
<style>body{{font-family:Segoe UI,Arial,sans-serif;margin:24px;color:#1f2328}}
table{{border-collapse:collapse;width:100%;margin:12px 0;font-size:14px}}
th,td{{border:1px solid #d0d7de;padding:6px 10px;text-align:left}} th{{background:#f6f8fa}} td.r{{text-align:right}}
.big{{font-size:20px;font-weight:700}} .badge{{color:#fff;padding:2px 8px;border-radius:10px;font-size:12px}}</style>
</head><body>
<h1>Nuvepro Moodle — Traceability &amp; Execution</h1>
<div>Generated {GENERATED} · manual cases from MoodleRegression_Testing.xlsx · scripts = Cucumber scenarios · results = target/cucumber.json</div>
<p class=big>Automated (created): {auto}/{tot} &nbsp;|&nbsp; Executed: {ex} &nbsp;|&nbsp; Passed: {ps}</p>
<h2>By area</h2>
<table><tr><th>Area</th><th>Sheet</th><th>Automated</th><th>Executed</th><th>Passed</th><th>Total</th><th>Coverage</th></tr>{summary}</table>
<h2>Automated cases → scenario → last result</h2>
<table><tr><th>Case</th><th>Manual title</th><th>Automated</th><th>Scenario</th><th>Last run</th></tr>{detail}</table>
</body></html>"""
    (OUT_DIR / "traceability_execution.html").write_text(doc, encoding="utf-8")

    print(f"Automated(created): {auto}/{tot} | Executed: {ex} | Passed: {ps}")
    print(f"Wrote {OUT_DIR / 'traceability_execution.csv'}")
    print(f"Wrote {OUT_DIR / 'traceability_execution.html'}")


if __name__ == "__main__":
    main()
